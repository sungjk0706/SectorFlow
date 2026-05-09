# -*- coding: utf-8 -*-
"""
증권사 API 명세 엑셀 파서
- 키움증권 OpenAPI 명세서 포맷 자동 인식
- 범용 포맷 (URL + Method + Name 컬럼 포함 시 자동 흡수)
반환: List[BrokerSpecRow]
"""
import re
import logging
from io import BytesIO
from typing import Any, Optional
from dataclasses import dataclass, field

import openpyxl

from app.core.broker_urls import KIWOOM_REST_REAL

logger = logging.getLogger(__name__)


@dataclass
class ParamDef:
    name: str
    type: str = "string"
    required: bool = False
    description: str = ""


@dataclass
class BrokerSpecRow:
    spec_name:       str
    method:          str          # GET | POST
    path:            str          # URL path (e.g. /uapi/domestic-stock/v1/...)
    base_url:        str = ""
    broker:          str = ""
    category:        str = ""
    tr_id:           str = ""
    description:     str = ""
    auth_type:       str = "bearer"
    extra_headers:   dict = field(default_factory=dict)
    request_schema:  list = field(default_factory=list)
    response_schema: list = field(default_factory=list)


# ── 유틸 ──────────────────────────────────────────────────────────────

def _clean(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _looks_like_path(s: str) -> bool:
    """'/uapi/...', '/stock/...', '/api/...' 처럼 슬래시로 시작하는 경로 감지"""
    return bool(re.match(r"^/[A-Za-z0-9_\-/.{}]+$", s.strip()))


def _extract_method(s: str) -> Optional[str]:
    s = s.upper()
    if "POST" in s:
        return "POST"
    if "GET" in s:
        return "GET"
    if "PUT" in s:
        return "PUT"
    if "DELETE" in s:
        return "DELETE"
    return None


def _cell_values(ws, row_idx: int) -> list[str]:
    return [_clean(ws.cell(row=row_idx, column=c).value) for c in range(1, ws.max_column + 2)]


# ── 키움 포맷 파서 ────────────────────────────────────────────────────

_KIWOOM_HEADER_KEYWORDS = {"tr코드", "tr code", "tr명", "url", "메서드", "api id", "api-id", "apiid"}


def _try_parse_kiwoom(ws) -> list[BrokerSpecRow]:
    """
    키움증권 OpenAPI 명세서 레이아웃 파싱.
    TR코드 / URL / 메서드 컬럼 기반.
    """
    header_row = None
    col_map: dict[str, int] = {}

    for r in range(1, min(20, ws.max_row + 1)):
        vals = _cell_values(ws, r)
        lower = [v.lower() for v in vals]
        hits = sum(1 for k in _KIWOOM_HEADER_KEYWORDS if k in lower)
        if hits >= 2:
            header_row = r
            for idx, v in enumerate(lower):
                col_map[v] = idx
            break

    if header_row is None:
        return []

    def col(keys: list[str]) -> Optional[int]:
        for k in keys:
            if k in col_map:
                return col_map[k]
        return None

    c_tr   = col(["api id", "api-id", "apiid", "tr코드", "tr code", "tr_id", "trid"])
    c_url  = col(["url"])
    c_meth = col(["메서드", "method", "http메서드"])
    c_name = col(["tr명", "api명", "name"])
    c_cat  = col(["분류", "category"])
    c_desc = col(["설명", "description"])

    if c_url is None:
        return []

    results: list[BrokerSpecRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = _cell_values(ws, r)
        url  = vals[c_url] if c_url is not None and c_url < len(vals) else ""
        if not url or not _looks_like_path(url):
            continue
        meth = _extract_method(vals[c_meth] if c_meth is not None and c_meth < len(vals) else "") or "GET"
        trid = vals[c_tr]   if c_tr   is not None and c_tr   < len(vals) else ""
        name = vals[c_name] if c_name is not None and c_name < len(vals) else (trid or url)
        cat  = vals[c_cat]  if c_cat  is not None and c_cat  < len(vals) else ""
        desc = vals[c_desc] if c_desc is not None and c_desc < len(vals) else ""

        extra: dict = {}
        if trid:
            extra["api-id"] = trid

        results.append(BrokerSpecRow(
            spec_name    = name or url,
            method       = meth,
            path         = url,
            base_url     = KIWOOM_REST_REAL,
            broker       = "kiwoom",
            category     = cat,
            tr_id        = trid,
            description  = desc,
            auth_type    = "bearer",
            extra_headers= extra,
        ))

    logger.info("키움 포맷 파싱 완료: %d 건", len(results))
    return results


# ── 범용 파서 ─────────────────────────────────────────────────────────

def _try_parse_generic(ws, broker: str) -> list[BrokerSpecRow]:
    """
    URL처럼 생긴 셀을 가진 행을 탐색.
    Method 컬럼이 있으면 사용, 없으면 GET 기본값.
    헤더 행 없이도 URL 패턴만으로 추출 가능.
    """
    results: list[BrokerSpecRow] = []

    # 1) 헤더 탐색
    header_row = None
    col_map: dict[str, int] = {}
    for r in range(1, min(15, ws.max_row + 1)):
        vals = _cell_values(ws, r)
        lower = [v.lower() for v in vals]
        if "url" in lower or "path" in lower:
            header_row = r
            for idx, v in enumerate(lower):
                col_map[v] = idx
            break

    c_url  = col_map.get("url") or col_map.get("path")
    c_meth = col_map.get("method") or col_map.get("메서드") or col_map.get("http method")
    c_name = col_map.get("name") or col_map.get("api명") or col_map.get("기능명")
    c_cat  = col_map.get("category") or col_map.get("분류")
    c_desc = col_map.get("description") or col_map.get("설명")

    start = (header_row + 1) if header_row else 1

    for r in range(start, ws.max_row + 1):
        vals = _cell_values(ws, r)
        if not vals:
            continue

        # URL을 컬럼 맵 기반 또는 전체 셀 스캔으로 탐색
        if c_url is not None and c_url < len(vals):
            url = vals[c_url]
        else:
            url = next((v for v in vals if _looks_like_path(v)), "")

        if not url or not _looks_like_path(url):
            continue

        meth_raw = vals[c_meth] if c_meth is not None and c_meth < len(vals) else ""
        meth     = _extract_method(meth_raw) or "POST"
        name     = vals[c_name] if c_name is not None and c_name < len(vals) else url
        cat      = vals[c_cat]  if c_cat  is not None and c_cat  < len(vals) else ""
        desc     = vals[c_desc] if c_desc is not None and c_desc < len(vals) else ""

        results.append(BrokerSpecRow(
            spec_name   = name or url,
            method      = meth,
            path        = url,
            broker      = broker,
            category    = cat,
            description = desc,
            auth_type   = "bearer",
        ))

    logger.info("범용 포맷 파싱 완료: %d 건", len(results))
    return results


# ── 공개 엔트리포인트 ─────────────────────────────────────────────────

def parse_broker_excel(
    data: bytes,
    broker: str = "custom",
    sheet_name: Optional[str] = None,
) -> list[BrokerSpecRow]:
    """
    엑셀 바이트를 받아 증권사 API 명세 목록을 반환.
    broker: 'kiwoom' | 'custom'
    sheet_name: None 이면 모든 시트를 파싱하여 합산
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {e}")

    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else list(wb.worksheets)

    all_rows: list[BrokerSpecRow] = []

    for ws in sheets:
        logger.info("시트 파싱 중: %s", ws.title)

        # 브로커별 전용 파서 우선 시도
        if broker == "kiwoom":
            rows = _try_parse_kiwoom(ws)
        else:
            rows = []

        # 전용 파서 실패 -> 범용 파서
        if not rows:
            rows = _try_parse_generic(ws, broker)

        # 브로커 정보 보정
        for row in rows:
            if not row.broker:
                row.broker = broker

        all_rows.extend(rows)

    wb.close()

    # 중복 제거 (같은 path + method)
    seen: set[str] = set()
    unique: list[BrokerSpecRow] = []
    for row in all_rows:
        key = f"{row.method}:{row.path}"
        if key not in seen:
            seen.add(key)
            unique.append(row)

    logger.info("최종 파싱 결과: %d 건 (중복 제거 후)", len(unique))
    return unique
